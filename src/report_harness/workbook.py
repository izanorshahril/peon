"""Excel workbook adapter for row-based report generation."""

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from .errors import WorkbookSchemaError
from .models import ProviderResponse, WorkbookRow


@dataclass(frozen=True)
class WorkbookColumns:
    evidence: str = "Evidence"
    answer: str = "Answer"
    justification: str = "Justification"


class ExcelWorkbook:
    """Read rows and update only the workbook's reserved output cells."""

    def __init__(
        self,
        path: Path | str,
        columns: WorkbookColumns | None = None,
        sheet_name: str | None = None,
    ) -> None:
        self.path = Path(path)
        self.columns = columns or WorkbookColumns()
        self._workbook = load_workbook(self.path, data_only=False)
        self._sheet = self._select_sheet(sheet_name)
        self._header_columns = self._find_required_columns()

    def iter_rows(self) -> Iterator[WorkbookRow]:
        header_row = self._sheet[1]
        headers = [cell.value for cell in header_row]
        for row_number in range(2, self._sheet.max_row + 1):
            cells = self._sheet[row_number]
            values = {
                str(headers[index]): cells[index].value
                for index in range(min(len(headers), len(cells)))
                if headers[index] is not None
            }
            if all(value is None for value in values.values()):
                continue
            evidence_value = cells[self._header_columns["evidence"] - 1].value
            evidence_reference = (
                str(evidence_value).strip() if evidence_value is not None else None
            )
            yield WorkbookRow(
                row_number=row_number,
                evidence_reference=evidence_reference,
                values=values,
            )

    def write_result(self, row_number: int, response: ProviderResponse) -> None:
        self._sheet.cell(
            row=row_number,
            column=self._header_columns["answer"],
            value=response.answer,
        )
        self._sheet.cell(
            row=row_number,
            column=self._header_columns["justification"],
            value=response.justification,
        )

    def save(self, path: Path | str | None = None) -> Path:
        output_path = Path(path) if path is not None else self.path
        self._workbook.save(output_path)
        return output_path

    def _select_sheet(self, sheet_name: str | None):
        if sheet_name is not None:
            try:
                return self._workbook[sheet_name]
            except KeyError as error:
                raise WorkbookSchemaError(
                    f"Workbook sheet not found: {sheet_name}"
                ) from error
        return self._workbook.active

    def _find_required_columns(self) -> dict[str, int]:
        requested = {
            "evidence": self.columns.evidence,
            "answer": self.columns.answer,
            "justification": self.columns.justification,
        }
        positions: dict[str, int] = {}
        for key, expected_header in requested.items():
            matches = [
                cell.column
                for cell in self._sheet[1]
                if self._normalise_header(cell.value)
                == self._normalise_header(expected_header)
            ]
            if not matches:
                raise WorkbookSchemaError(
                    f"Missing required workbook header: {expected_header}"
                )
            if len(matches) > 1:
                raise WorkbookSchemaError(
                    f"Ambiguous workbook header: {expected_header}"
                )
            positions[key] = matches[0]
        return positions

    @staticmethod
    def _normalise_header(value: object) -> str:
        return str(value).strip().casefold() if value is not None else ""