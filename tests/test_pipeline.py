from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from report_harness.models import EvidenceEnvelope, ProviderRequest, ProviderResponse
from report_harness.pipeline import ReportRunner
from report_harness.workbook import ExcelWorkbook


class StubEvidenceResolver:
    def resolve(self, row):
        return EvidenceEnvelope(
            reference=row.evidence_reference,
            media_type="image/png",
            content=b"image-bytes",
        )


class StubProvider:
    def justify(self, request: ProviderRequest) -> ProviderResponse:
        return ProviderResponse(
            answer=f"answer for {request.row_number}",
            justification=f"justification for {request.row_number}",
        )


def test_runner_processes_multiple_rows_and_preserves_unowned_cells(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.append(["Unrelated", "Justification", "Evidence", "Answer", "Formula"])
    sheet.append(["keep me", None, "one.png", None, "=1+1"])
    sheet.append(["also keep me", None, "two.png", None, "=2+2"])
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    workbook.save(input_path)

    runner = ReportRunner(
        workbook=ExcelWorkbook(input_path),
        evidence=StubEvidenceResolver(),
        provider=StubProvider(),
    )
    runner.run(output_path)

    result = load_workbook(output_path, data_only=False)
    result_sheet = result["Report"]
    assert result_sheet["B2"].value == "justification for 2"
    assert result_sheet["D2"].value == "answer for 2"
    assert result_sheet["B3"].value == "justification for 3"
    assert result_sheet["D3"].value == "answer for 3"
    assert result_sheet["A2"].value == "keep me"
    assert result_sheet["A3"].value == "also keep me"
    assert result_sheet["E2"].value == "=1+1"
    assert result_sheet["E3"].value == "=2+2"
    assert result_sheet["A2"].fill.fgColor.rgb == "00FFFF00"
