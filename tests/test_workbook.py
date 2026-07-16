from pathlib import Path

import pytest
from openpyxl import Workbook

from report_harness.errors import WorkbookSchemaError
from report_harness.workbook import ExcelWorkbook


def save_workbook(path: Path, headers: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(["evidence.png", None, None])
    workbook.save(path)


def test_missing_required_header_fails_before_rows_are_read(tmp_path: Path) -> None:
    path = tmp_path / "missing-header.xlsx"
    save_workbook(path, ["Evidence", "Answer", "Other"])

    with pytest.raises(WorkbookSchemaError, match="Justification"):
        ExcelWorkbook(path)


def test_duplicate_required_header_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-header.xlsx"
    save_workbook(path, ["Evidence", "Answer", "Justification", "evidence"])

    with pytest.raises(WorkbookSchemaError, match="Evidence"):
        ExcelWorkbook(path)


def test_reserved_columns_are_found_when_headers_move(tmp_path: Path) -> None:
    path = tmp_path / "moved-columns.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Item", "Answer", "Notes", "Evidence", "Justification"])
    sheet.append(["Item 1", None, "leave this", "evidence.png", None])
    workbook.save(path)

    adapter = ExcelWorkbook(path)
    row = next(adapter.iter_rows())

    assert row.evidence_reference == "evidence.png"
